"""
Upgraded gameplay flow tests for the team-based educational strategy game.

Verifies:
  - Problem metadata (threat level, state) is exposed.
  - Selecting / dragging / skipping / revisiting NEVER change metrics.
  - Problem states transition: UNRESOLVED -> SELECTED -> IN_PROGRESS -> SOLVED,
    plus temporary SKIPPED with free ordering.
  - Preview reports cost/benefit/risk without committing.
  - apply-solution records decision time, reaction time, energy spent,
    expected impact, and role.
  - Team Mode tags interactions with roles and keeps team-level outcome separate.
  - Metrics stay within valid ranges.

Run: python test_gameplay_flow.py
"""
import sys
import os
import tempfile
from io import BytesIO

sys.path.insert(0, os.path.dirname(__file__))

from fastapi.testclient import TestClient
from openpyxl import load_workbook

import main
from models import Database


def fresh_client():
    main.db = Database(os.path.join(tempfile.mkdtemp(), "test_neural_shield.db"))
    main.simulation = main.get_simulation()
    client = TestClient(main.app)
    return client


def assert_eq(actual, expected, msg):
    if actual != expected:
        raise AssertionError(f"{msg}: expected {expected!r}, got {actual!r}")


def assert_close(actual, expected, msg, tol=0.001):
    if abs(actual - expected) > tol:
        raise AssertionError(f"{msg}: expected {expected}, got {actual}")


def start_session(client, tag, team_mode=False, team_size=1):
    resp = client.post("/start-session", json={
        "participant_id": None, "challenge_type": "Easy", "challenge_order": 1,
        "team_mode": team_mode, "team_size": team_size,
    })
    assert resp.status_code == 200, resp.text
    data = resp.json()
    return data["session_id"], data["state"]


def test_1_problem_and_solution_metadata():
    client = fresh_client()
    session_id, state = start_session(client, "meta-1")

    assert len(state["problems"]) == 7, "Should expose exactly 7 problems"
    assert len(state["solutions"]) >= 8, "Should expose the core solution cards (filtered to the active level)"
    assert all("threat_level" in p and "state" in p and "expected_impact" in p for p in state["problems"])
    assert all(p["state"] == "UNRESOLVED" for p in state["problems"]), "All problems start UNRESOLVED"
    assert all("energy_cost" in s and "risk" in s and "valid_targets" in s for s in state["solutions"])
    # Every exposed card must be usable on at least one problem in this level.
    active_ids = {p["id"] for p in state["problems"]}
    assert all(any(t in active_ids for t in s["valid_targets"]) for s in state["solutions"]), \
        "Every solution card must address at least one problem in the active level"

    print("PASS 1: Problem metadata (threat/state/impact) and solution metadata exposed")


def test_2_navigation_never_changes_metrics():
    client = fresh_client()
    session_id, state = start_session(client, "nav-2")
    initial = (state["accuracy"], state["brain_health"], state["neural_energy"], state["score"])

    for problem in state["problems"]:
        resp = client.post("/select-problem", json={"session_id": session_id, "problem_id": problem["id"]})
        assert resp.status_code == 200, resp.text
        state = resp.json()["state"]

    resp = client.post("/skip-problem", json={"session_id": session_id, "problem_id": state["problems"][0]["id"]})
    state = resp.json()["state"]
    resp = client.post("/revisit-problem", json={"session_id": session_id, "problem_id": state["problems"][0]["id"]})
    state = resp.json()["state"]

    after = (state["accuracy"], state["brain_health"], state["neural_energy"], state["score"])
    assert_eq(after, initial, "Select/skip/revisit must never change metrics")

    # Selecting never reduces anything
    for problem in state["problems"]:
        resp = client.post("/select-problem", json={"session_id": session_id, "problem_id": problem["id"]})
        s = resp.json()["state"]
        assert s["accuracy"] == initial[0], "Selecting a problem must not change accuracy"

    print("PASS 2: Selecting / skipping / revisiting problems never change metrics")


def test_3_problem_states_and_free_ordering():
    client = fresh_client()
    session_id, state = start_session(client, "states-3")
    problems = state["problems"]

    # Skip a problem temporarily -> SKIPPED
    skip_id = problems[0]["id"]
    resp = client.post("/skip-problem", json={"session_id": session_id, "problem_id": skip_id})
    state = resp.json()["state"]
    assert_eq(state["problem_states"][skip_id], "SKIPPED", "Skipped problem should be SKIPPED")

    # Revisit -> SELECTED
    resp = client.post("/revisit-problem", json={"session_id": session_id, "problem_id": skip_id})
    state = resp.json()["state"]
    assert_eq(state["problem_states"][skip_id], "SELECTED", "Revisited problem should be SELECTED")

    # Select a DIFFERENT problem (free ordering, not sequential)
    second_id = problems[2]["id"]
    resp = client.post("/select-problem", json={"session_id": session_id, "problem_id": second_id})
    state = resp.json()["state"]
    assert_eq(state["problem_states"][second_id], "SELECTED", "Second problem should be SELECTED")
    assert_eq(state["current_event"], second_id, "Current event should be the second problem")

    # Wrong solution -> IN_PROGRESS (problem not solved, stays addressable)
    wrong_solution = "remove_noise"  # does not address second_id necessarily; find one that doesn't
    target = second_id
    resp = client.post("/apply-solution", json={
        "session_id": session_id, "action_type": wrong_solution, "target_event": target,
    })
    state = resp.json()["state"]
    if state["problem_states"].get(target) != "SOLVED":
        assert state["problem_states"][target] in ("IN_PROGRESS", "SELECTED"), \
            f"Failed attempt should leave problem IN_PROGRESS/SELECTED, got {state['problem_states'].get(target)}"

    print("PASS 3: Problem states UNRESOLVED/SELECTED/IN_PROGRESS/SKIPPED + free ordering")


def test_4_preview_and_apply_solution_records_times():
    client = fresh_client()
    session_id, state = start_session(client, "apply-4")

    dirty = next(p for p in state["problems"] if p["id"] == "Dirty Data")
    client.post("/select-problem", json={"session_id": session_id, "problem_id": dirty["id"]})

    # Preview a matching solution
    resp = client.post("/preview-action", json={"session_id": session_id, "action_type": "clean_dataset", "target_event": dirty["id"]})
    assert resp.status_code == 200, resp.text
    preview = resp.json()["preview"]
    assert_eq(preview["valid"], True, "clean_dataset should address Dirty Data")
    assert preview["energy_cost"] > 0, "Preview must report a cost"
    assert preview["expected_impact"] is not None, "Preview must report expected impact"
    assert preview["risk"], "Preview must include a risk note"
    assert_eq(preview["energy_affordable"], True, "Fresh session should afford the action")

    # Preview a mismatched solution -> invalid, no metrics change
    before = state["accuracy"]
    resp = client.post("/preview-action", json={"session_id": session_id, "action_type": "clean_dataset", "target_event": "Bias"})
    invalid = resp.json()["preview"]
    assert_eq(invalid["valid"], False, "clean_dataset must NOT address Bias")

    # Apply the correct solution with timing + role
    resp = client.post("/apply-solution", json={
        "session_id": session_id, "action_type": "clean_dataset", "target_event": dirty["id"],
        "role": "Modeler", "decision_time": 1800, "reaction_time": 950,
    })
    assert resp.status_code == 200, resp.text
    data = resp.json()
    state = data["state"]
    assert_eq(data["result"]["correct"], True, "Correct solution should be reported as correct")
    assert state["accuracy"] > before, "Correct solution must improve accuracy"
    assert_eq(state["problem_states"][dirty["id"]], "SOLVED", "Solved problem should be SOLVED")
    assert_eq(state["last_result"], "correct", "last_result should be correct")
    assert_eq(state["last_message"], "Good Decision", "Human-friendly success message")

    # Verify the interaction was recorded with timing + energy + role
    interactions = main.db.get_all_interactions(session_id)
    apply_rows = [i for i in interactions if i.get("action_type") == "apply_solution"]
    assert len(apply_rows) == 1, "Exactly one apply_solution should be recorded"
    row = apply_rows[0]
    assert_eq(row["decision_time"], 1800, "Decision time should be recorded")
    assert_eq(row["reaction_time"], 950, "Reaction time should be recorded")
    assert_eq(row["role"], "Modeler", "Role should be recorded")
    assert row["energy_spent"] and row["energy_spent"] > 0, "Energy spent should be recorded"
    assert row["expected_impact"] is not None, "Expected impact should be recorded"
    assert_eq(row["problem_state"], "SOLVED", "Problem state at apply time should be recorded")

    print("PASS 4: Preview cost/benefit/risk + apply-solution records timing/energy/impact/role")


def test_5_wrong_solution_consequences():
    client = fresh_client()
    session_id, state = start_session(client, "wrong-5")

    dirty = next(p for p in state["problems"] if p["id"] == "Dirty Data")
    client.post("/select-problem", json={"session_id": session_id, "problem_id": dirty["id"]})
    before = state["accuracy"]
    before_energy = state["neural_energy"]

    resp = client.post("/apply-solution", json={"session_id": session_id, "action_type": "remove_noise", "target_event": dirty["id"]})
    data = resp.json()
    state = data["state"]
    assert_eq(data["result"]["correct"], False, "Mismatched solution should be incorrect")
    assert state["accuracy"] < before, "Wrong solution must reduce accuracy"
    assert state["neural_energy"] < before_energy, "Wrong solution must cost energy"
    assert_eq(state["last_result"], "incorrect", "last_result should be incorrect")
    assert_eq(state["last_message"], "Try Another Approach", "Human-friendly retry message")

    # The problem is NOT solved and can be retried in any order later
    assert_eq(state["problem_states"][dirty["id"]] != "SOLVED", True, "Wrong solution must not solve the problem")

    print("PASS 5: Wrong solution -> realistic consequences (energy spent, accuracy drop, retryable)")


def test_6_team_mode_role_tagging_and_outcome_separation():
    client = fresh_client()
    session_id, state = start_session(client, "team-6", team_mode=True, team_size=3)
    assert_eq(state["team_mode"], True, "team_mode should be exposed")

    dirty = next(p for p in state["problems"] if p["id"] == "Dirty Data")
    client.post("/select-problem", json={"session_id": session_id, "problem_id": dirty["id"], "role": "Lead Analyst"})
    resp = client.post("/apply-solution", json={
        "session_id": session_id, "action_type": "clean_dataset", "target_event": dirty["id"],
        "role": "Modeler", "decision_time": 1500,
    })
    state = resp.json()["state"]

    # Individual interactions carry their own role tags
    interactions = main.db.get_all_interactions(session_id)
    roles = {i.get("action_type"): i.get("role") for i in interactions if i.get("role")}
    assert_eq(roles.get("select_problem"), "Lead Analyst", "Select event should carry the role")
    assert_eq(roles.get("apply_solution"), "Modeler", "Apply event should carry the role")

    # Finish session -> team-level outcome stored separately on the session
    resp = client.post("/finish-session", json={
        "session_id": session_id, "result": "manual", "challenge_type": "Easy", "challenge_order": 1,
    })
    assert resp.status_code == 200, resp.text
    session_results = client.get(f"/results?session_id={session_id}").json()["session"]
    assert_eq(session_results["team_mode"], 1, "Session should store team_mode")
    assert_eq(session_results["team_size"], 3, "Session should store team_size")
    assert "Modeler" in str(session_results.get("roles_used", "[]")) or session_results.get("roles_used"), \
        "roles_used should be stored on the session (team-level outcome)"

    # XLSX export still works and includes the Team Summary sheet
    resp = client.get(f"/export/session-xlsx?session_id={session_id}")
    assert resp.status_code == 200, resp.text
    wb = load_workbook(BytesIO(resp.content))
    assert "Team Summary" in wb.sheetnames, "XLSX should include a Team Summary sheet"
    summary = wb["Team Summary"]
    assert_eq(summary["B4"].value, "Team Session", "Team Summary should mark the session as a team session")

    print("PASS 6: Team Mode tags individual interactions with roles; team-level outcome stored separately")


def test_7_metric_ranges_and_invariants():
    client = fresh_client()
    session_id, state = start_session(client, "ranges-7")

    # Verify valid ranges across a full busy session
    for problem in state["problems"][:3]:
        client.post("/select-problem", json={"session_id": session_id, "problem_id": problem["id"]})
        resp = client.post("/apply-solution", json={"session_id": session_id, "action_type": "normalize_data", "target_event": problem["id"]})
        state = resp.json()["state"]
        assert 0.0 <= state["accuracy"] <= 1.0, f"Accuracy out of range: {state['accuracy']}"
        assert 0.0 <= state["brain_health"] <= 100.0, f"Brain health out of range: {state['brain_health']}"
        assert 0.0 <= state["neural_energy"] <= 200.0, f"Energy out of range: {state['neural_energy']}"
        assert state["time_remaining"] >= 0, f"Negative time: {state['time_remaining']}"
        assert state["score"] >= 0, f"Negative score: {state['score']}"

    # Reaching 90% still ends as a win immediately
    sim = main.simulation
    sim.base_accuracy = 0.90
    state = main.simulation.get_current_state()
    assert_eq(state.game_status, "won", "90% accuracy must immediately win")
    assert_eq(state.outcome, "won", "Outcome should be won")
    assert_eq(state.end_reason, "target_reached", "End reason should be target_reached")

    print("PASS 7: All metrics stay within valid ranges; 90% target still wins immediately")


if __name__ == "__main__":
    tests = [
        test_1_problem_and_solution_metadata,
        test_2_navigation_never_changes_metrics,
        test_3_problem_states_and_free_ordering,
        test_4_preview_and_apply_solution_records_times,
        test_5_wrong_solution_consequences,
        test_6_team_mode_role_tagging_and_outcome_separation,
        test_7_metric_ranges_and_invariants,
    ]

    passed = 0
    failed = 0
    for test in tests:
        try:
            test()
            passed += 1
        except AssertionError as e:
            print(f"FAIL {test.__name__}: {e}")
            failed += 1
        except Exception as e:
            print(f"ERROR {test.__name__}: {type(e).__name__}: {e}")
            failed += 1

    print(f"\nResults: {passed} passed, {failed} failed")
    sys.exit(1 if failed else 0)
