"""
Tests for SOLO/TEAM game mode separation, role switching, and role bonuses.

Run: python test_modes.py

Verifies:
  - start-session records game_mode = solo / team on the session row
  - /set-role records a role_switch interaction and never changes metrics
  - Team Mode role bonuses are small, deterministic, and only apply with a role
  - Solo Mode maths are untouched (no role -> identical rewards)
  - XLSX export contains a Game Mode row and team fields for team sessions only
"""
import sys
import os
import tempfile
from io import BytesIO

sys.path.insert(0, os.path.dirname(__file__))

from fastapi.testclient import TestClient
from openpyxl import load_workbook

import main
from models import Database


def fresh_client():
    main.db = Database(os.path.join(tempfile.mkdtemp(), "test_modes.db"))
    main.simulation = main.get_simulation()
    client = TestClient(main.app)
    return client


def assert_eq(actual, expected, msg):
    if actual != expected:
        raise AssertionError(f"{msg}: expected {expected!r}, got {actual!r}")


def start(client, game_mode="solo", team_size=1):
    team_mode = game_mode == "team"
    resp = client.post("/start-session", json={
        "participant_id": None, "challenge_type": "Easy", "challenge_order": 1,
        "team_mode": team_mode, "team_size": team_size, "game_mode": game_mode,
    })
    assert resp.status_code == 200, resp.text
    data = resp.json()
    return data["session_id"], data["state"]


def test_1_game_mode_recorded():
    client = fresh_client()

    sid_solo, state = start(client, "solo")
    assert_eq(state["game_mode"], "solo", "State should expose game_mode solo")
    assert_eq(state["team_mode"], False, "Solo must not set team_mode")
    assert_eq(state["active_role"], None, "Solo has no active role")

    sid_team, state = start(client, "team", team_size=3)
    assert_eq(state["game_mode"], "team", "State should expose game_mode team")
    assert_eq(state["team_mode"], True, "Team mode should set team_mode")
    assert_eq(state["team_size"], 3, "Team size should be recorded")

    # Verify DB row
    solo_row = main.db.get_session_results(sid_solo)
    team_row = main.db.get_session_results(sid_team)
    assert_eq(solo_row.get("game_mode"), "solo", "Session row should store game_mode solo")
    assert_eq(team_row.get("game_mode"), "team", "Session row should store game_mode team")

    print("PASS modes_1: game_mode recorded in state + session row")


def test_2_set_role_never_changes_metrics():
    client = fresh_client()
    sid, state = start(client, "team", team_size=2)
    before = (state["accuracy"], state["brain_health"], state["neural_energy"], state["score"])

    resp = client.post("/set-role", json={"session_id": sid, "role": "Team Lead"})
    assert resp.status_code == 200, resp.text
    state = resp.json()["state"]
    assert_eq(state["active_role"], "Team Lead", "Active role should be set")

    after = (state["accuracy"], state["brain_health"], state["neural_energy"], state["score"])
    assert_eq(after, before, "Switching roles must never change metrics")

    # Invalid role is ignored
    resp = client.post("/set-role", json={"session_id": sid, "role": "Not A Role"})
    state = resp.json()["state"]
    assert_eq(state["active_role"], "Team Lead", "Invalid role must be ignored")

    # role_switch interaction recorded
    interactions = main.db.get_all_interactions(sid)
    switches = [i for i in interactions if i.get("action_type") == "role_switch"]
    assert_eq(len(switches), 1, "One role_switch should be recorded")
    assert_eq(switches[0]["role"], "Team Lead", "Role should be on the interaction")

    print("PASS modes_2: /set-role changes nothing but the active role; interaction recorded")


def test_3_role_bonus_deterministic_and_solo_unchanged():
    client = fresh_client()

    # Solo: correct clean_dataset on Dirty Data gives the base reward (0.5 -> 0.536)
    sid_solo, _ = start(client, "solo")
    client.post("/select-problem", json={"session_id": sid_solo, "problem_id": "Dirty Data"})
    r = client.post("/apply-solution", json={"session_id": sid_solo, "action_type": "clean_dataset", "target_event": "Dirty Data"})
    assert_eq(r.json()["result"]["correct"], True, "clean_dataset should correct Dirty Data")
    acc_solo = r.json()["state"]["accuracy"]
    assert_eq(acc_solo, 0.536, "Solo clean_dataset on Dirty Data must stay 0.5->0.536")

    # Team with Data Analyst: same action must give strictly more reward, deterministically
    sid_team, state = start(client, "team", team_size=2)
    client.post("/set-role", json={"session_id": sid_team, "role": "Data Analyst"})
    client.post("/select-problem", json={"session_id": sid_team, "problem_id": "Dirty Data"})
    r = client.post("/apply-solution", json={"session_id": sid_team, "action_type": "clean_dataset", "target_event": "Dirty Data"})
    acc_team = r.json()["state"]["accuracy"]
    assert acc_team > acc_solo, f"Data Analyst bonus should improve reward (solo={acc_solo}, team={acc_team})"
    # Deterministic: same role, same action, same event -> same result
    sid_team2, _ = start(client, "team", team_size=2)
    client.post("/set-role", json={"session_id": sid_team2, "role": "Data Analyst"})
    client.post("/select-problem", json={"session_id": sid_team2, "problem_id": "Dirty Data"})
    r2 = client.post("/apply-solution", json={"session_id": sid_team2, "action_type": "clean_dataset", "target_event": "Dirty Data"})
    assert_eq(r2.json()["state"]["accuracy"], acc_team, "Role bonus must be deterministic")

    # Security Analyst: wrong solution penalty is smaller (mild)
    sid_sec, _ = start(client, "team", team_size=2)
    client.post("/set-role", json={"session_id": sid_sec, "role": "Security Analyst"})
    client.post("/select-problem", json={"session_id": sid_sec, "problem_id": "Dirty Data"})
    r3 = client.post("/apply-solution", json={"session_id": sid_sec, "action_type": "remove_noise", "target_event": "Dirty Data"})
    acc_sec = r3.json()["state"]["accuracy"]

    sid_plain, _ = start(client, "team", team_size=2)  # team WITHOUT a role -> no multiplier
    client.post("/select-problem", json={"session_id": sid_plain, "problem_id": "Dirty Data"})
    r4 = client.post("/apply-solution", json={"session_id": sid_plain, "action_type": "remove_noise", "target_event": "Dirty Data"})
    acc_plain = r4.json()["state"]["accuracy"]
    assert acc_sec >= acc_plain, f"Security Analyst should soften the penalty (sec={acc_sec}, plain={acc_plain})"
    assert acc_sec > 0, "Accuracy must stay >= 0"

    print("PASS modes_3: role bonuses small, deterministic; solo maths unchanged")


def test_4_xlsx_game_mode_field():
    client = fresh_client()

    sid_solo, state = start(client, "solo")
    client.post("/select-problem", json={"session_id": sid_solo, "problem_id": "Dirty Data"})
    client.post("/apply-solution", json={"session_id": sid_solo, "action_type": "clean_dataset", "target_event": "Dirty Data"})
    client.post("/finish-session", json={"session_id": sid_solo, "result": "timeout", "game_mode": "solo"})
    resp = client.get(f"/export/session-xlsx?session_id={sid_solo}")
    assert resp.status_code == 200, resp.text
    wb = load_workbook(BytesIO(resp.content))
    summary = wb["Research Summary"]
    # Find the Game Mode row (appended after the Outcome section)
    game_mode_cell = None
    for row in range(1, 40):
        if summary.cell(row=row, column=1).value == "Game Mode":
            game_mode_cell = summary.cell(row=row, column=2).value
            break
    assert_eq(game_mode_cell, "Solo", "XLSX should show Game Mode = Solo for solo sessions")

    sid_team, state = start(client, "team", team_size=3)
    client.post("/set-role", json={"session_id": sid_team, "role": "Data Analyst"})
    client.post("/select-problem", json={"session_id": sid_team, "problem_id": "Dirty Data", "role": "Data Analyst"})
    client.post("/apply-solution", json={"session_id": sid_team, "action_type": "clean_dataset", "target_event": "Dirty Data", "role": "Data Analyst"})
    client.post("/finish-session", json={"session_id": sid_team, "result": "manual", "game_mode": "team"})
    resp = client.get(f"/export/session-xlsx?session_id={sid_team}")
    wb = load_workbook(BytesIO(resp.content))

    # Team Summary should show Team Session + game_mode + team size + role fields
    assert "Team Summary" in wb.sheetnames
    team_sheet = wb["Team Summary"]
    values_by_label = {}
    for row in range(1, 20):
        label = team_sheet.cell(row=row, column=1).value
        val = team_sheet.cell(row=row, column=2).value
        if label:
            values_by_label[label] = val
    assert_eq(values_by_label.get("Mode"), "Team Session", "Team Summary should say Team Session")
    assert_eq(values_by_label.get("Game Mode"), "Team", "Team Summary should show Game Mode = Team")
    assert "Data Analyst" in str(values_by_label.get("Roles Used", "")), "Roles Used should contain the role"

    # Action Log has Role populated for team
    action_sheet = wb["Action Log"]
    headers = [action_sheet.cell(row=1, column=c).value for c in range(1, 20)]
    assert "Role" in headers, "Action Log should have Role column"

    print("PASS modes_4: XLSX carries game_mode; team rows carry role data")


if __name__ == "__main__":
    tests = [
        test_1_game_mode_recorded,
        test_2_set_role_never_changes_metrics,
        test_3_role_bonus_deterministic_and_solo_unchanged,
        test_4_xlsx_game_mode_field,
    ]
    passed = 0
    failed = 0
    for test in tests:
        try:
            test()
            passed += 1
        except AssertionError as e:
            print(f"FAIL {test.__name__}: {e}")
            failed += 1
        except Exception as e:
            print(f"ERROR {test.__name__}: {type(e).__name__}: {e}")
            failed += 1
    print(f"\nResults: {passed} passed, {failed} failed")
    sys.exit(1 if failed else 0)
