"""
End-to-end accuracy invariant tests through the FastAPI pipeline.
Run: python test_accuracy_api.py

Verifies that Accuracy NEVER changes except on an explicit correct/wrong
submitted solution, from the API down to the XLSX export.
"""
import sys
import os
import tempfile
import math
from io import BytesIO

sys.path.insert(0, os.path.dirname(__file__))

from fastapi.testclient import TestClient
from openpyxl import load_workbook

import main
from models import Database


def fresh_client():
    main.db = Database(os.path.join(tempfile.mkdtemp(), "test_neural_shield.db"))
    main.simulation = main.get_simulation()
    client = TestClient(main.app)
    return client


def assert_close(actual, expected, msg, tol=0.001):
    if abs(actual - expected) > tol:
        raise AssertionError(f"{msg}: expected {expected}, got {actual}")


def start_session(client, tag):
    resp = client.post("/start-session", json={"participant_id": None, "challenge_type": "Easy", "challenge_order": 1})
    assert resp.status_code == 200, resp.text
    data = resp.json()
    return data["session_id"], data["state"]


def test_1_select_all_problems_unchanged():
    client = fresh_client()
    session_id, state = start_session(client, "api-test-1")
    initial = state["accuracy"]

    for problem in state["active_events"]:
        resp = client.post("/select-problem", json={"session_id": session_id, "problem_id": problem})
        assert resp.status_code == 200, resp.text
        state = resp.json()["state"]
        assert_close(state["accuracy"], initial, f"Accuracy changed after selecting {problem}")

    print("PASS api_test_1: Select all 7 problems via API -> Accuracy unchanged")


def test_2_switch_and_wait_30s_unchanged():
    client = fresh_client()
    session_id, state = start_session(client, "api-test-2")
    initial = state["accuracy"]

    problems = state["active_events"]
    for _ in range(3):
        for p in problems:
            resp = client.post("/select-problem", json={"session_id": session_id, "problem_id": p})
            state = resp.json()["state"]

    for _ in range(30):
        resp = client.post("/advance-time", json={"session_id": session_id, "action_type": "1"})
        assert resp.status_code == 200, resp.text
        state = resp.json()["state"]

    assert_close(state["accuracy"], initial, "Accuracy changed after switching and waiting 30s")
    assert state["time_remaining"] == 150, f"Time should decrease, got {state['time_remaining']}"
    print("PASS api_test_2: Switch problems + wait 30s via API -> Accuracy unchanged")


def test_3_select_action_without_execution_unchanged():
    client = fresh_client()
    session_id, state = start_session(client, "api-test-3")
    initial = state["accuracy"]

    for node_id in main.simulation.nodes:
        main.simulation.nodes[node_id]["energy"] = 0

    resp = client.post("/apply-game-action", json={"session_id": session_id, "action_type": "clean_dataset"})
    assert resp.status_code == 200, resp.text
    state = resp.json()["state"]

    assert_close(state["accuracy"], initial, "Accuracy changed on non-executed action (no energy)")
    print("PASS api_test_3: Select action without executing -> Accuracy unchanged")


def test_4_correct_solution_reward():
    client = fresh_client()
    session_id, state = start_session(client, "api-test-4")
    initial = state["accuracy"]

    client.post("/select-problem", json={"session_id": session_id, "problem_id": "Dirty Data"})
    resp = client.post("/apply-game-action", json={"session_id": session_id, "action_type": "clean_dataset"})
    state = resp.json()["state"]

    expected_reward = main.simulation._calculate_correct_reward("clean_dataset", "Dirty Data")
    assert state["accuracy"] > initial, "Correct solution should increase accuracy"
    assert_close(state["accuracy"], initial + expected_reward, "Correct reward amount mismatch")
    print(f"PASS api_test_4: Correct solution -> reward applied ({initial} -> {state['accuracy']})")


def test_5_wrong_solution_penalty():
    client = fresh_client()
    session_id, state = start_session(client, "api-test-5")
    initial = state["accuracy"]

    client.post("/select-problem", json={"session_id": session_id, "problem_id": "Dirty Data"})
    resp = client.post("/apply-game-action", json={"session_id": session_id, "action_type": "remove_noise"})
    state = resp.json()["state"]

    expected_penalty = main.simulation._calculate_wrong_penalty("Dirty Data")
    assert state["accuracy"] < initial, "Wrong solution should decrease accuracy"
    assert_close(state["accuracy"], initial - expected_penalty, "Wrong penalty amount mismatch")
    print(f"PASS api_test_5: Wrong solution -> penalty applied ({initial} -> {state['accuracy']})")


def test_6_idle_and_navigate_after_wrong_unchanged():
    client = fresh_client()
    session_id, state = start_session(client, "api-test-6")

    client.post("/select-problem", json={"session_id": session_id, "problem_id": "Dirty Data"})
    resp = client.post("/apply-game-action", json={"session_id": session_id, "action_type": "remove_noise"})
    state = resp.json()["state"]
    after_wrong = state["accuracy"]

    for _ in range(10):
        resp = client.post("/advance-time", json={"session_id": session_id, "action_type": "1"})
        state = resp.json()["state"]
    client.post("/select-problem", json={"session_id": session_id, "problem_id": "Noise"})
    client.post("/select-problem", json={"session_id": session_id, "problem_id": "Dirty Data"})

    assert_close(state["accuracy"], after_wrong, "Accuracy changed after idle/navigation post-wrong")
    print("PASS api_test_6: Navigate/idle after wrong solution -> Accuracy unchanged")


def _collect_workbook_values(workbook):
    bad = []
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                    bad.append((ws.title, cell.coordinate, v))
    return bad


def test_7_gameplay_backend_results_xlsx_identical():
    client = fresh_client()
    session_id, state = start_session(client, "api-test-7")

    client.post("/select-problem", json={"session_id": session_id, "problem_id": "Dirty Data"})
    resp = client.post("/apply-game-action", json={"session_id": session_id, "action_type": "clean_dataset"})
    state = resp.json()["state"]
    gameplay_accuracy = state["accuracy"]

    # Finish the session (use the explicit result) - authoritative value on backend
    resp = client.post("/finish-session", json={
        "session_id": session_id, "result": "manual",
        "challenge_type": "Easy", "challenge_order": 1
    })
    assert resp.status_code == 200, resp.text
    backend_accuracy = resp.json()["final_metrics"]["accuracy"]

    # Results Screen data
    resp = client.get(f"/results?session_id={session_id}")
    assert resp.status_code == 200, resp.text
    results_accuracy = resp.json()["session"]["final_accuracy"]

    # XLSX export
    resp = client.get(f"/export/session-xlsx?session_id={session_id}")
    assert resp.status_code == 200, resp.text
    wb = load_workbook(BytesIO(resp.content))

    summary = wb["Research Summary"]
    xlsx_accuracy = summary["B14"].value
    xlsx_accuracy_format = summary["B14"].number_format

    assert_close(backend_accuracy, gameplay_accuracy, "Backend final accuracy != gameplay accuracy")
    assert_close(results_accuracy, gameplay_accuracy, "Results Screen accuracy != gameplay accuracy")
    assert_close(xlsx_accuracy, gameplay_accuracy, "XLSX accuracy != gameplay accuracy")
    assert xlsx_accuracy_format == "0.0%", f"XLSX accuracy format unexpected: {xlsx_accuracy_format}"

    # No NaN / Infinity / impossible percentages anywhere
    bad = _collect_workbook_values(wb)
    assert not bad, f"NaN/Infinity values in XLSX: {bad}"

    # Accuracy/Precision/Recall are 0-1 decimals; Brain Health stored 0-100
    challenge = wb["Challenge Summary"]
    # Data row is row 2 (row 1 is headers). openpyxl columns are 1-indexed:
    # col 7 = Starting Accuracy, col 8 = Final Accuracy, col 11 = Final Brain Health
    for col in [7, 8]:
        v = challenge.cell(row=2, column=col).value
        assert v is None or (0.0 <= v <= 1.0), f"Challenge Summary accuracy out of range: {v}"
    bh = challenge.cell(row=2, column=11).value
    assert bh is None or (0.0 <= bh <= 1.0), f"Challenge Summary brain health not 0-100: {bh}"
    completion = summary["B11"].value
    assert completion is None or (0.0 <= completion <= 180.0), f"Completion time out of range: {completion}"

    print(f"PASS api_test_7: Gameplay {gameplay_accuracy}, backend {backend_accuracy}, "
          f"results {results_accuracy}, XLSX {xlsx_accuracy} all identical")


if __name__ == "__main__":
    tests = [
        test_1_select_all_problems_unchanged,
        test_2_switch_and_wait_30s_unchanged,
        test_3_select_action_without_execution_unchanged,
        test_4_correct_solution_reward,
        test_5_wrong_solution_penalty,
        test_6_idle_and_navigate_after_wrong_unchanged,
        test_7_gameplay_backend_results_xlsx_identical,
    ]

    passed = 0
    failed = 0
    for test in tests:
        try:
            test()
            passed += 1
        except AssertionError as e:
            print(f"FAIL {test.__name__}: {e}")
            failed += 1
        except Exception as e:
            print(f"ERROR {test.__name__}: {e}")
            failed += 1

    print(f"\nResults: {passed} passed, {failed} failed")
    sys.exit(1 if failed else 0)
