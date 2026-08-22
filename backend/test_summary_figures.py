"""
Performance Summary figure integrity tests for AI Brain Lab.

The frontend advances the game clock once per second and records navigation
(select/skip/revisit). Those rows must NEVER inflate the performance summary.
This suite replays a realistic session (clock polls + navigation + correct and
incorrect solutions) and verifies every summary figure:

  1. Total/Correct/Incorrect actions count ONLY real solution decisions.
  2. Total Events is the mission problem count (7), not the interaction row count.
  3. Success Rate and Event Success Rate are mathematically consistent.
  4. Every summary figure is identical across finish-session, /results and XLSX.
  5. Brain Health stats are reported as 0-100% (never 9500%).
  6. advance_time rows never leak seconds/energy into energy fields.

Run: python test_summary_figures.py
"""
import sys
import os
import math
import tempfile
from io import BytesIO

sys.path.insert(0, os.path.dirname(__file__))

from fastapi.testclient import TestClient
from openpyxl import load_workbook

import main
from models import Database


def fresh_client():
    main.db = Database(os.path.join(tempfile.mkdtemp(), "test_summary.db"))
    main.simulation = main.get_simulation()
    return TestClient(main.app)


def assert_eq(actual, expected, msg):
    if actual != expected:
        raise AssertionError(f"{msg}: expected {expected!r}, got {actual!r}")


def assert_close(actual, expected, msg, tol=0.001):
    if abs(actual - expected) > tol:
        raise AssertionError(f"{msg}: expected {expected}, got {actual}")


def play_realistic_session(client, session_id):
    """Replay what the frontend does: clock polls + navigation + solutions."""
    # 45 seconds of clock polling before the first decision.
    for _ in range(45):
        client.post("/advance-time", json={"session_id": session_id, "action_type": "1"})

    # 3 correct solutions.
    for problem, action in [
        ("Dirty Data", "clean_dataset"),
        ("Missing Values", "normalize_data"),
        ("Class Imbalance", "balance_dataset"),
    ]:
        client.post("/select-problem", json={"session_id": session_id, "problem_id": problem})
        client.post("/apply-solution", json={
            "session_id": session_id, "action_type": action,
            "target_event": problem, "role": "Data Analyst",
            "decision_time": 3000, "reaction_time": 800,
        })

    # 2 incorrect solutions (valid energy, wrong card).
    for problem, action in [
        ("Noise", "clean_dataset"),
        ("Data Drift", "remove_noise"),
    ]:
        client.post("/select-problem", json={"session_id": session_id, "problem_id": problem})
        client.post("/apply-solution", json={
            "session_id": session_id, "action_type": action,
            "target_event": problem, "role": "Data Analyst",
            "decision_time": 5000, "reaction_time": 1200,
        })

    # A skipped problem + revisit (navigation only).
    client.post("/select-problem", json={"session_id": session_id, "problem_id": "Bias"})
    client.post("/skip-problem", json={"session_id": session_id, "problem_id": "Bias", "role": "Data Analyst"})
    client.post("/revisit-problem", json={"session_id": session_id, "problem_id": "Bias", "role": "Data Analyst"})

    # 30 more seconds of clock polling.
    for _ in range(30):
        client.post("/advance-time", json={"session_id": session_id, "action_type": "1"})


def test_1_action_totals_exclude_clock_and_navigation():
    client = fresh_client()
    resp = client.post("/start-session", json={"participant_id": None, "challenge_type": "Easy", "challenge_order": 1})
    session_id = resp.json()["session_id"]

    play_realistic_session(client, session_id)

    resp = client.post("/finish-session", json={
        "session_id": session_id, "result": "manual",
        "challenge_type": "Easy", "challenge_order": 1,
    })
    metrics = resp.json()["final_metrics"]

    # Exactly 5 solution cards were committed: 3 solved + 2 wrong.
    assert_eq(metrics["total_actions"], 5, "total_actions must count only solution decisions")
    assert_eq(metrics["correct_actions"], 3, "correct_actions mismatch")
    assert_eq(metrics["wrong_actions"], 2, "wrong_actions mismatch")
    assert_eq(metrics["total_actions"], metrics["correct_actions"] + metrics["wrong_actions"],
              "action consistency: total == correct + wrong")

    # Events are mission problems (7), independent of interaction count.
    assert_eq(metrics["total_events"], 7, "total_events must be the mission size")
    assert_eq(metrics["events_solved"], 3, "events_solved mismatch")

    # Navigation stats stay separate and intact.
    assert metrics["skipped_count"] >= 1, "skipped_count should include the skip"
    assert metrics["problems_selected"] >= 5, "problems_selected should include the selects"

    # The stored results row must match the finish payload exactly.
    results = client.get(f"/results?session_id={session_id}").json()["session"]
    assert_eq(results["total_actions"], 5, "results total_actions mismatch")
    assert_eq(results["correct_actions"], 3, "results correct_actions mismatch")
    assert_eq(results["wrong_actions"], 2, "results wrong_actions mismatch")
    assert_eq(results["total_events"], 7, "results total_events mismatch")
    print("PASS 1: actions/events figures correct (5 attempts, 3 correct, 2 wrong, 7 events)")


def test_2_xlsx_summary_matches_stored_metrics():
    client = fresh_client()
    resp = client.post("/start-session", json={"participant_id": None, "challenge_type": "Easy", "challenge_order": 2})
    session_id = resp.json()["session_id"]

    play_realistic_session(client, session_id)

    client.post("/finish-session", json={
        "session_id": session_id, "result": "manual",
        "challenge_type": "Easy", "challenge_order": 2,
    })

    resp = client.get(f"/export/session-xlsx?session_id={session_id}")
    assert resp.status_code == 200, resp.text
    wb = load_workbook(BytesIO(resp.content))
    summary = wb["Research Summary"]

    # Decision Performance block (rows 20-23).
    assert_eq(summary["B20"].value, 5, "XLSX Total Actions")
    assert_eq(summary["B21"].value, 3, "XLSX Correct Actions")
    assert_eq(summary["B22"].value, 2, "XLSX Incorrect Actions")
    assert_close(summary["B23"].value, 0.6, "XLSX Success Rate (3/5)")

    # Events block (rows 26-29).
    assert_eq(summary["B26"].value, 7, "XLSX Total Events")
    assert_eq(summary["B27"].value, 3, "XLSX Events Solved")
    assert_eq(summary["B28"].value, 4, "XLSX Events Ignored (7-3)")
    assert_close(summary["B29"].value, 3 / 7, "XLSX Event Success Rate (3/7)")

    # Research Metrics: Brain Health row (3) mean/min/max (cols B-D) are 0-100%,
    # stored as decimals (0.5 = 50%), never >1.0.
    metrics_sheet = wb["Research Metrics"]
    bh_values = [metrics_sheet.cell(row=3, column=c).value for c in range(2, 5)]
    for v in bh_values:
        assert v is not None and 0.0 <= v <= 1.0, f"Brain Health stat out of range: {v}"

    print("PASS 2: XLSX summary figures correct and Research Metrics brain health sane")


def test_3_advance_time_does_not_leak_into_energy_fields():
    client = fresh_client()
    resp = client.post("/start-session", json={"participant_id": None, "challenge_type": "Easy", "challenge_order": 3})
    session_id = resp.json()["session_id"]

    play_realistic_session(client, session_id)

    interactions = client.get(f"/results?session_id={session_id}").json()["interactions"]
    clock_rows = [i for i in interactions if i.get("action_type") == "advance_time"]
    solution_rows = [i for i in interactions if i.get("action_type") == "apply_solution"]

    assert len(clock_rows) == 75, f"expected 75 advance_time rows, got {len(clock_rows)}"

    # Clock rows must not carry seconds as energy_allocated or a bogus energy_spent.
    for row in clock_rows:
        assert row.get("energy_allocated") is None, f"advance_time leaked seconds into energy_allocated: {row.get('energy_allocated')}"
        assert row.get("energy_spent") in (None, 0), f"advance_time bogus energy_spent: {row.get('energy_spent')}"

    # Solution rows must record an energy_spent (not empty).
    for row in solution_rows:
        assert row.get("energy_spent") is not None, "apply_solution must record energy_spent"
        assert row.get("energy_spent") >= 0, "energy_spent must be non-negative"

    print("PASS 3: clock rows stay out of energy fields; solutions record energy_spent")


def test_4_no_nan_inf_or_impossible_values_anywhere():
    client = fresh_client()
    resp = client.post("/start-session", json={"participant_id": None, "challenge_type": "Easy", "challenge_order": 4})
    session_id = resp.json()["session_id"]

    play_realistic_session(client, session_id)
    client.post("/finish-session", json={
        "session_id": session_id, "result": "manual",
        "challenge_type": "Easy", "challenge_order": 4,
    })

    resp = client.get(f"/export/session-xlsx?session_id={session_id}")
    wb = load_workbook(BytesIO(resp.content))
    bad = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                    bad.append((ws.title, cell.coordinate))
    assert not bad, f"NaN/Infinity in workbook: {bad}"
    print("PASS 4: no NaN / Infinity in any workbook cell")


if __name__ == "__main__":
    tests = [
        test_1_action_totals_exclude_clock_and_navigation,
        test_2_xlsx_summary_matches_stored_metrics,
        test_3_advance_time_does_not_leak_into_energy_fields,
        test_4_no_nan_inf_or_impossible_values_anywhere,
    ]

    passed = 0
    failed = 0
    for test in tests:
        try:
            test()
            passed += 1
        except AssertionError as e:
            print(f"FAIL {test.__name__}: {e}")
            failed += 1
        except Exception as e:
            import traceback
            traceback.print_exc()
            print(f"ERROR {test.__name__}: {e}")
            failed += 1

    print(f"\nResults: {passed} passed, {failed} failed")
    sys.exit(1 if failed else 0)
